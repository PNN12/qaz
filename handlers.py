import asyncio

from aiogram.types import Message, BotCommand
from aiogram.dispatcher.storage import FSMContext
from aiogram.utils.markdown import hbold, hunderline

from main import dp, bot, path, title
from keyboards import *
from states import *
from words import *
from function import *

# pip install pandas openpyxl xlsxwriter xlrd
from openpyxl import load_workbook

"""                                  ____          _      _____              _ 
                                    | __ )   ___  | |_   | ____|__  __  ___ | |
                                    |  _ \  / _ \ | __|  |  _|  \ \/ / / _ \| |
                                    | |_) || (_) || |_   | |___  >  < |  __/| |
                                    |____/  \___/  \__|  |_____|/_/\_\ \___||_|                                      
"""


# ------------------------------------ Base handlers (start, cancel) ---------------------------------------

@dp.message_handler(commands=["start"], state="*")
async def command_start(message: Message, state: FSMContext) -> None:
    await bot.set_my_commands([BotCommand("start", "Активировать бот"),
                               BotCommand("download_exel", "Загрузить готовый Exel файл"),
                               BotCommand("find_values", "Вывод всех значений категории")])
    username = f"@{message.from_user.full_name}" if message.from_user.username is None else f"@{message.from_user.username}"
    await message.answer_photo(photo=open(f"photos/start_photo.png", "rb"), reply_markup=get_start_kb(),
                               caption=hbold(f"👋 Привет {username}\n") + START_DESCRIPTION, protect_content=True)
    await message.delete()
    await state.finish()


@dp.message_handler(text=CANCEL, state="*")
async def command_cancel(message: Message, state: FSMContext) -> None:
    if state is None:
        pass
    await message.answer(hunderline(CANCEL_WRITING_DATA), reply_markup=get_start_kb())
    await message.delete()
    await state.finish()


# ---------------------------------------- Menu handlers ----------------------------------------------

#                                                                                                         Download Exel
@dp.message_handler(text=[DOWNLOAD_EXEL, "/download_exel"], state="*")
async def download_exel(message: Message, state: FSMContext) -> None:
    await state.finish()
    await message.answer(text=hbold(SEND_EXEL_FILE), reply_markup=get_cancel_kb())
    await message.delete()
    await DownloadExel.download_exel.set()

    # Хендлер проверки: является ли message документом
    @dp.message_handler(lambda message: not message.document, state=DownloadExel.download_exel)
    async def check_exel_file(message: Message) -> None:
        await message.answer(text=hunderline(IS_NOT_EXEL_FILE))

    @dp.message_handler(content_types=["document"], state=DownloadExel.download_exel)
    async def output_download_exel(message: Message, state: FSMContext) -> None:
        async with state.proxy() as data:
            data["download_exel"] = message.document.file_name
            name_document = data["download_exel"]

        global path, title
        await message.document.download(destination_file=f"files/{name_document}")
        path = name_document
        title = load_workbook(f"files/{path}").sheetnames[0]
        await message.answer_document(document=open(f"files/{path}", "rb"), reply_markup=get_start_kb(),
                                      caption=f"{hbold(LOADED_EXEL_FILE)}\n\n")
        await state.finish()

#                                                                                                           Find Values
@dp.message_handler(text=[FIND_VALUES, "/find_values"], state="*")
async def find_values(message: Message, state: FSMContext) -> None:
    try:
        global path, title
        await state.finish()
        listbook = load_workbook(f'files/{path}')[title]
        out_category = ""
        for idx, val in enumerate(get_categories(listbook=listbook, category=True)):
            out_category += f"{idx + 1}) {val}\n"
        await message.answer(text=f"{hbold('Категории:')}\n{out_category}\n{SHOW_SELECTED_CATEGORY}",
                             reply_markup=get_cancel_kb())
        await FindValues.find_values.set()
    except Exception:
        await message.answer(text=hunderline(NOT_SELECTED_FILE_EXCEL), reply_markup=get_start_kb())

    @dp.message_handler(state=FindValues.find_values)
    async def output_find_values(message: Message, state: FSMContext) -> None:
        async with state.proxy() as data:
            data["find_values"] = message.text
            text_find = data["find_values"]
        global path, title
        listbook = load_workbook(f"files/{path}")[title]

        try:
            out_text, out_list = "", []
            categories: list = get_categories(listbook=listbook, text=text_find)  # type -> list: [a, b, c]
            values: dict = get_values(listbook=listbook, text=text_find)  # type -> dict: {row: [a, b, c]}

            if not categories and not values:
                raise Exception("Message text is empty")

            elif categories:
                for row, column in enumerate(categories):
                    out_text += f"{row + 1}) {column}\n"
                    if row % 100 == 0:
                        out_list.append(out_text)
                        out_text = ""
                else:
                    out_list.append(out_text)
            else:
                for row, list_value in values.items():
                    value = lambda num: list_value[num] if list_value[num] is not None else 0

                    ordering = (value(8), value(9), value(10), value(11), value(12), value(13), value(14), value(15))
                    min_price = max(ordering) if min(ordering) == 0 else min(ordering)

                    out_list.append(f"<u>Строка: №{row}\n</u>"
                                    f"<i>Раздел:</i> <b>{list_value[1]}\n</b>"
                                    f"<i>Полная модель:</i> <b>{list_value[2]}\n</b>"
                                    f"<i>Mодель:</i> <b>{list_value[3]}\n</b>"
                                    f"<i>Память:</i> <b>{list_value[4]}\n</b>"
                                    f"<i>Цвет:</i> <b>{list_value[5]}\n</b>"
                                    f"<i>Тип:</i> <b>{list_value[6]}\n</b>"
                                    f"<i>G5:</i> <b>{list_value[7]}\n</b>"
                                    f"<b>Цена от Abdulloh: {value(8)}\n</b>"
                                    f"<b>Цена от Axror: {value(9)}\n</b>"
                                    f"<b>Цена от Alisher: {value(10)}\n</b>"
                                    f"<b>Цена от Abror: {value(11)}\n</b>"
                                    f"<b>Цена от Abusaxiy: {value(12)}\n</b>"
                                    f"<b>Цена от Aks33: {value(13)}\n</b>"
                                    f"<b>Цена от ApplePro: {value(14)}\n</b>"
                                    f"<b>Цена от B5: {value(15)}\n</b>"
                                    f"<i>Минимальная стоимость: {min_price}</i>")
            for i in out_list:
                await message.answer(text=i)
                await asyncio.sleep(1.5)

        except Exception as error:
            if str(error) == "Message text is empty":
                await message.answer(text=hunderline(ERROR_MESSAGE_EMPTY))  # Error №1: Message text is empty
            elif str(error).find("Flood control exceeded.") >= 0:
                await message.answer(text=hunderline(ERROR_FLOOD))  # Error №2: Flood control exceeded.
            else:
                await message.answer(text=hunderline(ERROR_ANOTHER))  # Error №3: Another errors
        finally:
            await state.finish()
            await find_values(message, state)

